"""
Build management_report.xlsx from the cleaned retail data.

Input:  clean/retail_sales_2025_clean.csv  (produced by scripts/clean_retail_sales.py)
Output: management_report.xlsx  with three sheets:
    - "Summary":   total revenue, order count, average order value (formatted block).
    - "By Region": revenue pivot (region x category) with a total row, currency
                   formatting, and the top-performing region highlighted.
    - "Monthly":   monthly revenue with an embedded line chart.

All numbers here are reproducible by rerunning this script.

Usage:
    python scripts/build_management_report.py
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

CLEAN_PATH = "clean/retail_sales_2025_clean.csv"
OUT_PATH = "management_report.xlsx"

# ── Reusable styles ─────────────────────────────────────────────────────────
BOLD = Font(bold=True)
TITLE = Font(bold=True, size=14)
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")          # dark blue header
HIGHLIGHT_FILL = PatternFill("solid", fgColor="FFF2CC")       # soft yellow highlight
TOTAL_FILL = PatternFill("solid", fgColor="E2EFDA")           # soft green total row
CURRENCY = '"$"#,##0.00'                                       # currency number format
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def autofit(ws, widths: dict[int, int]) -> None:
    """Set column widths from a {col_index: char_width} map (openpyxl can't
    truly auto-fit, so we size from the longest rendered value per column)."""
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width + 2


def measure(rows: list[list]) -> dict[int, int]:
    """Longest string length per column index (1-based) across the given rows."""
    widths: dict[int, int] = {}
    for row in rows:
        for i, val in enumerate(row, start=1):
            widths[i] = max(widths.get(i, 0), len(str(val)))
    return widths


def build_summary(wb, total_revenue, order_count, aov) -> None:
    ws = wb.active
    ws.title = "Summary"

    ws["A1"] = "Retail Sales 2025 — Management Report"
    ws["A1"].font = TITLE

    # Label / value block with bold labels.
    block = [
        ("Total Revenue", total_revenue, CURRENCY),
        ("Order Count", order_count, "#,##0"),
        ("Average Order Value", aov, CURRENCY),
    ]
    for r, (label, value, fmt) in enumerate(block, start=3):
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = BOLD
        vc = ws.cell(row=r, column=2, value=value)
        vc.number_format = fmt
        vc.alignment = Alignment(horizontal="right")

    autofit(ws, {1: len("Average Order Value"), 2: 14})
    # Freeze everything above the metric block so the title/labels stay in view.
    ws.freeze_panes = "A3"


def build_by_region(wb, pivot: pd.DataFrame, top_region: str) -> None:
    ws = wb.create_sheet("By Region")
    categories = list(pivot.columns)                 # category cols + trailing "Total"
    header = ["Region"] + categories

    # Header row.
    for c, name in enumerate(header, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    # Data rows (one per region), highlighting the top-performing region.
    for r, (region, values) in enumerate(pivot.iterrows(), start=2):
        rc = ws.cell(row=r, column=1, value=region)
        rc.font = BOLD
        rc.border = BORDER
        for c, cat in enumerate(categories, start=2):
            cell = ws.cell(row=r, column=c, value=float(values[cat]))
            cell.number_format = CURRENCY
            cell.border = BORDER
        if region == top_region:
            for c in range(1, len(header) + 1):
                ws.cell(row=r, column=c).fill = HIGHLIGHT_FILL

    # Total row: column-wise sums across all regions.
    total_row = len(pivot) + 2
    tc = ws.cell(row=total_row, column=1, value="Total")
    tc.font = BOLD
    tc.fill = TOTAL_FILL
    tc.border = BORDER
    for c, cat in enumerate(categories, start=2):
        cell = ws.cell(row=total_row, column=c, value=float(pivot[cat].sum()))
        cell.number_format = CURRENCY
        cell.font = BOLD
        cell.fill = TOTAL_FILL
        cell.border = BORDER

    # Size columns from header + region names.
    widths = {1: max(len("Region"), *(len(str(x)) for x in pivot.index))}
    for c, name in enumerate(categories, start=2):
        widths[c] = max(len(str(name)), 13)          # 13 fits "$1,234,567.89"
    autofit(ws, widths)

    # Freeze the header row and the Region label column.
    ws.freeze_panes = "B2"

    # Note under the table calling out the highlighted region.
    note_row = total_row + 2
    ws.cell(row=note_row, column=1,
            value=f"Highlighted: {top_region} (top-performing region by total revenue)").font = \
        Font(italic=True)


def build_monthly(wb, monthly: pd.DataFrame) -> None:
    ws = wb.create_sheet("Monthly")

    # Header.
    for c, name in enumerate(["Month", "Revenue"], start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Data rows.
    for r, row in enumerate(monthly.itertuples(index=False), start=2):
        ws.cell(row=r, column=1, value=row.month_label)
        rev = ws.cell(row=r, column=2, value=float(row.revenue))
        rev.number_format = CURRENCY

    last = len(monthly) + 1

    # Embedded line chart of monthly revenue.
    chart = LineChart()
    chart.title = "Monthly Revenue — 2025"
    chart.y_axis.title = "Revenue ($)"
    chart.x_axis.title = "Month"
    chart.height = 8
    chart.width = 18
    data = Reference(ws, min_col=2, min_row=1, max_row=last)   # include header as series name
    cats = Reference(ws, min_col=1, min_row=2, max_row=last)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    ws.add_chart(chart, "D2")

    autofit(ws, {1: 8, 2: 14})
    ws.freeze_panes = "A2"


def main() -> None:
    # ── Load cleaned data ────────────────────────────────────────────────────
    df = pd.read_csv(CLEAN_PATH, parse_dates=["order_date"])

    # ── Summary metrics ──────────────────────────────────────────────────────
    total_revenue = float(df["revenue ($)"].sum())
    order_count = int(df["order_id"].nunique())
    aov = total_revenue / order_count

    # ── Region x category revenue pivot, with a per-region Total column ──────
    pivot = df.pivot_table(index="region", columns="category",
                           values="revenue ($)", aggfunc="sum", fill_value=0)
    pivot["Total"] = pivot.sum(axis=1)
    pivot = pivot.sort_values("Total", ascending=False)   # best region on top
    top_region = pivot["Total"].idxmax()

    # ── Monthly revenue (calendar month) ─────────────────────────────────────
    m = df.copy()
    m["month"] = m["order_date"].dt.to_period("M")
    monthly = (m.groupby("month")["revenue ($)"].sum()
                 .reset_index()
                 .rename(columns={"revenue ($)": "revenue"}))
    monthly["month_label"] = monthly["month"].dt.strftime("%Y-%m")
    monthly = monthly[["month_label", "revenue"]]

    # ── Build workbook ───────────────────────────────────────────────────────
    wb = Workbook()
    build_summary(wb, total_revenue, order_count, aov)
    build_by_region(wb, pivot, top_region)
    build_monthly(wb, monthly)
    wb.save(OUT_PATH)

    # ── Console echo of the reported figures ─────────────────────────────────
    print("=== REPORTED FIGURES ===")
    print(f"Total revenue:        ${total_revenue:,.2f}")
    print(f"Order count:          {order_count:,}")
    print(f"Average order value:  ${aov:,.2f}")
    print(f"Top-performing region: {top_region} (${pivot.loc[top_region, 'Total']:,.2f})")
    print(f"Months covered:       {len(monthly)}")
    print(f"Saved workbook:       {OUT_PATH}")


if __name__ == "__main__":
    main()
