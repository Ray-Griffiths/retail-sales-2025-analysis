"""
Rebuild management_report.xlsx and add an executive Dashboard sheet at the front.

Input:  clean/retail_sales_2025_clean.csv  (from scripts/clean_retail_sales.py)
Output: management_report.xlsx  with sheets:
    - "Dashboard": KPI cards + four charts (region, category, monthly trend,
                   region x category mix). This is the visual management view.
    - "Summary", "By Region", "Monthly": the underlying detail (also the
      "table view" that backs the dashboard charts).

WHY rebuild instead of append: openpyxl drops existing charts when it loads and
re-saves a workbook, so we regenerate every sheet in one fresh Workbook. The
detail sheets are built by reusing the functions in build_management_report.py.

Chart design follows the dataviz skill: validated categorical palette, single-hue
magnitude bars, direct labels on charts whose colors need contrast relief, legends
for multi-series charts only.

Usage:
    python scripts/build_dashboards.py
"""

import os
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Reuse the detail-sheet builders and shared styles from the report script.
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from build_management_report import (  # noqa: E402
    BOLD,
    BORDER,
    CURRENCY,
    HEADER_FILL,
    HEADER_FONT,
    TITLE,
    build_by_region,
    build_monthly,
    build_summary,
)

CLEAN_PATH = "clean/retail_sales_2025_clean.csv"
AUDIT_PATH = "clean/cleaning_audit.csv"
OUT_PATH = "management_report.xlsx"

# ── Validated categorical palette (dataviz skill, light surface) ─────────────
# Slots 1-5, used for the 5 product categories in fixed order. CVD-safe (worst
# adjacent ΔE 24.2). Contrast relief for aqua/yellow = direct labels + table view.
CAT_COLORS = ["2A78D6", "1BAF7A", "EDA100", "008300", "4A3AA7"]
BRAND_BLUE = "2A78D6"          # single-hue magnitude (region bar, monthly line)

# KPI card accent fills (decorative, coordinated with the palette).
CARD_FILLS = ["305496", "1BAF7A", "4A3AA7", "EB6834"]

# Chrome / ink from the palette's chart-chrome table.
INK_PRIMARY = "0B0B0B"
MUTED = "898781"
BANNER_FILL = "1F3864"         # deep blue title banner
WHITE = "FFFFFF"


def style_range_fill(ws, cell_range: str, fill: PatternFill) -> None:
    """Apply one fill across every cell in a rectangular range (for card blocks)."""
    for row in ws[cell_range]:
        for cell in row:
            cell.fill = fill


def make_kpi_card(ws, top: int, left: int, label: str, value: str, fill_hex: str) -> None:
    """Draw a 3-col x 4-row KPI card: coloured block, big white value, small label."""
    right = left + 2                       # card spans 3 columns
    fill = PatternFill("solid", fgColor=fill_hex)
    lc, rc = get_column_letter(left), get_column_letter(right)

    # Colour the whole block first so merged + surrounding cells share the fill.
    style_range_fill(ws, f"{lc}{top}:{rc}{top + 3}", fill)

    # Label band (top row of the card).
    ws.merge_cells(f"{lc}{top}:{rc}{top}")
    lab = ws.cell(row=top, column=left, value=label.upper())
    lab.font = Font(bold=True, size=10, color="D9E1F2")
    lab.alignment = Alignment(horizontal="center", vertical="center")

    # Value (big, spans the remaining three rows).
    ws.merge_cells(f"{lc}{top + 1}:{rc}{top + 3}")
    val = ws.cell(row=top + 1, column=left, value=value)
    val.font = Font(bold=True, size=22, color=WHITE)
    val.alignment = Alignment(horizontal="center", vertical="center")


def _style_axes(chart) -> None:
    """Quiet, recessive axes: muted labels, no vertical gridlines."""
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.majorGridlines = None
    for ax in (chart.x_axis, chart.y_axis):
        ax.spPr = None  # keep default hairline; avoids heavy borders


def chart_revenue_by_region(byregion_ws) -> BarChart:
    """Horizontal bar, single blue hue, sorted (sheet is pre-sorted desc),
    largest on top, value labels for precision."""
    chart = BarChart()
    chart.type = "bar"
    chart.title = "Revenue by Region"
    chart.height, chart.width = 8.5, 15
    # Total column is G (col 7); regions in col A rows 2-7 (6 regions).
    data = Reference(byregion_ws, min_col=7, min_row=1, max_row=7)
    cats = Reference(byregion_ws, min_col=1, min_row=2, max_row=7)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None                                   # single series
    chart.series[0].graphicalProperties.solidFill = BRAND_BLUE
    chart.series[0].graphicalProperties.line.noFill = True
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.numFmt = '"$"#,##0'
    chart.y_axis.scaling.orientation = "maxMin"           # largest region on top
    _style_axes(chart)
    return chart


def chart_revenue_by_category(byregion_ws) -> DoughnutChart:
    """Doughnut of category share, one categorical hue per slice, % labels."""
    chart = DoughnutChart()
    chart.title = "Revenue Share by Category"
    chart.height, chart.width = 8.5, 15
    # Category totals live in the "Total" row (row 8), cols B-F; headers row 1.
    data = Reference(byregion_ws, min_col=2, max_col=6, min_row=8, max_row=8)
    cats = Reference(byregion_ws, min_col=2, max_col=6, min_row=1, max_row=1)
    chart.add_data(data, from_rows=True)
    chart.set_categories(cats)
    chart.holeSize = 55
    # Colour each slice with its category hue.
    series = chart.series[0]
    for i, color in enumerate(CAT_COLORS):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color
        pt.graphicalProperties.line.solidFill = "FFFFFF"  # 2px surface gap
        series.data_points.append(pt)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.numFmt = "0.0%"
    return chart


def chart_monthly_trend(monthly_ws, last_row: int) -> LineChart:
    """Single-series line, 2px blue, markers — revenue over the 12 months."""
    chart = LineChart()
    chart.title = "Monthly Revenue Trend"
    chart.height, chart.width = 8.5, 15
    data = Reference(monthly_ws, min_col=2, min_row=1, max_row=last_row)
    cats = Reference(monthly_ws, min_col=1, min_row=2, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = BRAND_BLUE
    s.graphicalProperties.line.width = 25400                # 2pt in EMU
    s.marker = Marker(symbol="circle", size=6)
    s.smooth = False
    _style_axes(chart)
    return chart


def chart_region_category_mix(byregion_ws) -> BarChart:
    """Stacked column: category mix within each region. Legend + category hues.
    Relief for low-contrast hues is the 'By Region' table view + the legend."""
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = "Category Mix by Region"
    chart.height, chart.width = 8.5, 15
    # Five category columns B-F, headers in row 1; regions on the x-axis rows 2-7.
    data = Reference(byregion_ws, min_col=2, max_col=6, min_row=1, max_row=7)
    cats = Reference(byregion_ws, min_col=1, min_row=2, max_row=7)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    for i, color in enumerate(CAT_COLORS):
        chart.series[i].graphicalProperties.solidFill = color
        chart.series[i].graphicalProperties.line.solidFill = "FFFFFF"  # segment gap
    chart.legend.position = "b"
    _style_axes(chart)
    return chart


def build_data_quality(wb, audit_df: pd.DataFrame) -> None:
    """Add an auditor-facing 'Data Quality' sheet from the cleaning audit trail.

    Renders one row per cleaning action with the row counts affected, plus a
    reconciliation line so the raw -> clean row count can be traced end to end.
    The numbers come straight from clean/cleaning_audit.csv (emitted by the
    cleaning script), so the sheet cannot drift from what actually ran."""
    ws = wb.create_sheet("Data Quality")          # appended after the detail sheets
    ws.sheet_view.showGridLines = False

    # ── Title + provenance ───────────────────────────────────────────────────
    ws["A1"] = "Data Quality — Cleaning Audit Trail"
    ws["A1"].font = TITLE
    ws["A2"] = ("Source: retail_sales_2025.xlsx (read-only) · produced by "
                "scripts/clean_retail_sales.py → clean/cleaning_audit.csv · "
                "rendered by scripts/build_dashboards.py")
    ws["A2"].font = Font(italic=True, size=9, color=MUTED)

    # ── Reconciliation callout (raw − duplicates − returns = clean) ──────────
    def _affected(keyword: str) -> int:
        hit = audit_df[audit_df["action"].str.contains(keyword, case=False)]
        return int(pd.to_numeric(hit["rows_affected"], errors="coerce").fillna(0).sum())

    raw_rows = int(pd.to_numeric(audit_df["rows_out"], errors="coerce").iloc[0])
    clean_rows = int(pd.to_numeric(audit_df["rows_out"], errors="coerce").iloc[-1])
    dupes, returns = _affected("duplicate"), _affected("return")
    recon = (f"Reconciliation:  {raw_rows:,} raw  −  {dupes:,} duplicates  −  "
             f"{returns:,} returns  =  {clean_rows:,} clean rows")
    ws["A4"] = recon
    ws["A4"].font = Font(bold=True, size=11, color="1F3864")
    ws.merge_cells("A4:G4")

    # ── Table header ─────────────────────────────────────────────────────────
    headers = ["Step", "Cleaning Action", "Basis (rule)", "Rows In",
               "Rows Affected", "Rows Out", "Detail"]
    header_row = 6
    for c, name in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    # ── Table body ───────────────────────────────────────────────────────────
    num_cols = {"rows_in", "rows_affected", "rows_out"}
    field_by_col = ["step", "action", "basis", "rows_in",
                    "rows_affected", "rows_out", "detail"]
    zebra = PatternFill("solid", fgColor="F2F5FA")     # subtle alternating shade

    for i, (_, rec) in enumerate(audit_df.iterrows()):
        r = header_row + 1 + i
        for c, field in enumerate(field_by_col, start=1):
            raw_val = rec[field]
            # Blank out NaN / empty; render count columns as integers.
            if pd.isna(raw_val) or raw_val == "":
                value = None
            elif field in num_cols:
                value = int(float(raw_val))
            else:
                value = raw_val
            cell = ws.cell(row=r, column=c, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(
                horizontal="center" if field in num_cols or field == "step" else "left",
                vertical="center",
                wrap_text=(field == "detail"))
            if field in num_cols and value is not None:
                cell.number_format = "#,##0"
            if field == "action":
                cell.font = BOLD
        if i % 2 == 1:                                 # shade odd rows
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = zebra

    # ── Column widths + freeze the header ────────────────────────────────────
    widths = {1: 6, 2: 28, 3: 24, 4: 9, 5: 14, 6: 9, 7: 62}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = f"A{header_row + 1}"


def build_dashboard(wb, kpis: dict, monthly_last_row: int) -> None:
    """Create the front 'Dashboard' sheet: banner, KPI cards, four charts."""
    ws = wb.create_sheet("Dashboard", 0)          # index 0 = first/front sheet
    ws.sheet_view.showGridLines = False           # clean canvas

    # Even canvas columns A-P for a tidy card/chart grid.
    for c in range(1, 17):
        ws.column_dimensions[get_column_letter(c)].width = 9.5

    # ── Title banner ─────────────────────────────────────────────────────────
    ws.merge_cells("A1:P2")
    style_range_fill(ws, "A1:P2", PatternFill("solid", fgColor=BANNER_FILL))
    t = ws.cell(row=1, column=1, value="Retail Sales 2025  —  Executive Dashboard")
    t.font = Font(bold=True, size=20, color=WHITE)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22

    sub = ws.cell(row=3, column=1,
                  value="Full-year 2025 · net of returns · figures reproduced by scripts/build_dashboards.py")
    sub.font = Font(italic=True, size=9, color=MUTED)

    # ── KPI cards (row 5-8), four across with one-column gaps ─────────────────
    cards = [
        ("Total Revenue", kpis["total_revenue"]),
        ("Total Orders", kpis["order_count"]),
        ("Avg Order Value", kpis["aov"]),
        (f"Top Region · {kpis['top_region']}", kpis["top_region_rev"]),
    ]
    left_cols = [1, 5, 9, 13]                     # A, E, I, M
    for (label, value), left, fill in zip(cards, left_cols, CARD_FILLS):
        make_kpi_card(ws, top=5, left=left, label=label, value=value, fill_hex=fill)
    for r in range(5, 9):
        ws.row_dimensions[r].height = 20

    # ── Charts (2 x 2 grid) referencing the detail sheets ────────────────────
    byregion_ws = wb["By Region"]
    monthly_ws = wb["Monthly"]
    ws.add_chart(chart_revenue_by_region(byregion_ws), "A10")
    ws.add_chart(chart_revenue_by_category(byregion_ws), "I10")
    ws.add_chart(chart_monthly_trend(monthly_ws, monthly_last_row), "A28")
    ws.add_chart(chart_region_category_mix(byregion_ws), "I28")


def main() -> None:
    # ── Load cleaned data ────────────────────────────────────────────────────
    df = pd.read_csv(CLEAN_PATH, parse_dates=["order_date"])

    # ── Metrics for the detail sheets ────────────────────────────────────────
    total_revenue = float(df["revenue ($)"].sum())
    order_count = int(df["order_id"].nunique())
    aov = total_revenue / order_count

    pivot = df.pivot_table(index="region", columns="category",
                           values="revenue ($)", aggfunc="sum", fill_value=0)
    pivot["Total"] = pivot.sum(axis=1)
    pivot = pivot.sort_values("Total", ascending=False)
    top_region = pivot["Total"].idxmax()
    top_region_rev = float(pivot.loc[top_region, "Total"])

    m = df.copy()
    m["month"] = m["order_date"].dt.to_period("M")
    monthly = (m.groupby("month")["revenue ($)"].sum()
                 .reset_index().rename(columns={"revenue ($)": "revenue"}))
    monthly["month_label"] = monthly["month"].dt.strftime("%Y-%m")
    monthly = monthly[["month_label", "revenue"]]
    monthly_last_row = len(monthly) + 1                # +1 for header row

    # ── Load the cleaning audit trail (must exist — run the clean script) ────
    if not os.path.exists(AUDIT_PATH):
        raise SystemExit(f"Missing {AUDIT_PATH}. Run: python scripts/clean_retail_sales.py")
    audit_df = pd.read_csv(AUDIT_PATH, dtype=str)   # read as text; coerce per-column

    # ── Build workbook: detail sheets first (charts reference them) ──────────
    wb = Workbook()
    build_summary(wb, total_revenue, order_count, aov)
    build_by_region(wb, pivot, top_region)
    build_monthly(wb, monthly)
    build_data_quality(wb, audit_df)

    # KPI values formatted for the cards.
    kpis = {
        "total_revenue": f"${total_revenue:,.0f}",
        "order_count": f"{order_count:,}",
        "aov": f"${aov:,.0f}",
        "top_region": top_region,
        "top_region_rev": f"${top_region_rev:,.0f}",
    }
    build_dashboard(wb, kpis, monthly_last_row)
    wb.save(OUT_PATH)

    # ── Console echo ─────────────────────────────────────────────────────────
    print("=== DASHBOARD BUILT ===")
    print("sheets:", wb.sheetnames)
    print(f"Total revenue: ${total_revenue:,.2f} | Orders: {order_count:,} | "
          f"AOV: ${aov:,.2f} | Top region: {top_region} (${top_region_rev:,.0f})")
    print(f"Saved: {OUT_PATH}")


if __name__ == "__main__":
    main()
